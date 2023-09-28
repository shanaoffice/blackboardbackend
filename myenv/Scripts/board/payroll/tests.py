# import openpyxl
# from reportlab.lib.pagesizes import letter
# from reportlab.pdfgen import canvas

# def convert_excel_to_pdf(input_excel_path, output_pdf_path):
#     wb = openpyxl.load_workbook(input_excel_path)
#     ws = wb.active

#     pdf = canvas.Canvas(output_pdf_path, pagesize=letter)

#     x, y = 100, 700  # Starting position

#     for row in ws.iter_rows():
#         for cell in row:
#             pdf.drawString(x, y, str(cell.value))
#             x += 100  # Adjust as needed for column width
#         y -= 15  # Adjust as needed for row height
#         x = 100

#     pdf.save()

# # Provide paths to your input Excel file and desired output PDF file
# input_excel_path = "C:/Users/Admin/Desktop/Blackboard Data Sheets/Book15.xlsx"
# output_pdf_path = "C:/Users/Admin/Desktop/Blackboard Data Sheets/output.pdf"

# convert_excel_to_pdf(input_excel_path, output_pdf_path)

# import  jpype     
# import  asposecells     
# jpype.startJVM() 
# from asposecells.api import Workbook
# workbook = Workbook("C:/Users/Admin/Desktop/Blackboard Data Sheets/Book15.xlsx")
# workbook.save("C:/Users/Admin/Desktop/Blackboard Data Sheets/output1.pdf")
# jpype.shutdownJVM()

from openpyxl import Workbook
from openpyxl import load_workbook

# Create a new workbook and select the active sheet
wb =load_workbook("D:/Varjinth/Salary-Slip.xlsx")
sheet = wb.active

merged_start_col, merged_end_col = 1, 3
merged_start_row, merged_end_row = 3, 3


# Calculate the new value
new_value = "New Value"

# Get the merged cell's previous value
previous_value = sheet.cell(row=merged_start_row, column=merged_start_col).value

combined_value = f"{previous_value}, {new_value}"
sheet["A3"]= sheet["A3"].value + "saddd"

# Iterate through merged cells and update them
# for row in sheet.iter_rows(min_col=merged_start_col, max_col=merged_end_col,
#                             min_row=merged_start_row, max_row=merged_end_row):
#     for cell in row:
#         print(cell.value) 
        # cell.value=123
# Save the workbook
# wb.save('modified_example.xlsx')
# # Add data to cells
# sheet['A1'] = 'Name'
# sheet['B1'] = 'Age'

# sheet['A2'] = 'Alice'
# sheet['B2'] = 25

# sheet['A3'] = 'Bob'
# sheet['B3'] = 30

# # Apply some formatting
# from openpyxl.styles import Font

# bold_font = Font(bold=True)
# sheet['A1'].font = bold_font
# sheet['B1'].font = bold_font

# Save the workbook
wb.save('example.xlsx')
